import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Changelog"

# ── Palette ────────────────────────────────────────────────────────────────
BLACK      = "FF18181B"
WHITE      = "FFFFFFFF"
ACCENT     = "FFFECA76"   # golden
LIGHT_GREY = "FFF4F4F5"
MID_GREY   = "FFE4E4E7"
TEXT_MUTED = "FF71717A"
GREEN_BG   = "FFD1FAE5"
GREEN_FG   = "FF065F46"
BLUE_BG    = "FFDBEAFE"
BLUE_FG    = "FF1E40AF"
PURPLE_BG  = "FFEDE9FE"
PURPLE_FG  = "FF5B21B6"
RED_BG     = "FFFEE2E2"
RED_FG     = "FF991B1B"
AMBER_BG   = "FFFEF3C7"
AMBER_FG   = "FF92400E"

def fill(hex_colour):
    return PatternFill("solid", fgColor=hex_colour)

def border():
    thin = Side(style="thin", color=MID_GREY)
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def badge_style(bg, fg):
    return {"fill": fill(bg), "font": Font(bold=True, size=9, color=fg)}

# ── Category badge colours ─────────────────────────────────────────────────
CATEGORY_STYLES = {
    "Schema / Migration": badge_style(PURPLE_BG, PURPLE_FG),
    "Backend Model":      badge_style(BLUE_BG,   BLUE_FG),
    "Backend Route":      badge_style(BLUE_BG,   BLUE_FG),
    "Backend Service":    badge_style(BLUE_BG,   BLUE_FG),
    "Frontend Page":      badge_style(GREEN_BG,  GREEN_FG),
    "Frontend Config":    badge_style(GREEN_BG,  GREEN_FG),
    "Frontend Style":     badge_style(GREEN_BG,  GREEN_FG),
    "Bug Fix":            badge_style(RED_BG,    RED_FG),
    "Documentation":      badge_style(AMBER_BG,  AMBER_FG),
}

# ── Data ───────────────────────────────────────────────────────────────────
VERSION_COLOURS = {
    "1.0.0": "FFFDE68A",
    "1.1.0": "FFBFDBFE",
    "1.2.0": "FFBBF7D0",
    "1.3.0": "FFFECACA",
}

rows = [
    # Version, Timestamp (AEST), Commit short, Category, File(s) changed, Objective / description
    (
        "1.0.0",
        "30 May 2026  12:13 pm AEST",
        "5d05ca8",
        "Schema / Migration",
        "049_business_units.sql\n050_accounts_contacts.sql\n051_lead_pipeline.sql\n052_leads.sql",
        "Phase 1 — Introduce core CRM schema: Business Units with member roles, Accounts & Contacts, pipeline stages per BU, and the full Leads lifecycle (estimates, routing rules, activity log).",
    ),
    (
        "1.1.0",
        "30 May 2026  2:28 pm AEST",
        "9141cee",
        "Backend Model",
        "Account.js\nBusinessUnit.js\nLead.js\nPipelineStage.js",
        "Phase 2 — CRUD model layer for Business Units, Accounts/Contacts, Pipeline Stages, and Leads. Includes lead routing engine (dot-notation field matching) and inbound lead ingestion API endpoint.",
    ),
    (
        "1.1.0",
        "30 May 2026  2:28 pm AEST",
        "9141cee",
        "Backend Route",
        "accountRoutes.js\nbusinessUnitRoutes.js\nleadRoutes.js\napiV1Leads.js\napiV1.js\nplatformRouter.js",
        "Phase 2 — REST routes for BU management, accounts, leads pipeline, and the public lead-ingest API. Webhook events fired on lead.created / lead.won / lead.lost.",
    ),
    (
        "1.1.0",
        "30 May 2026  2:28 pm AEST",
        "9141cee",
        "Backend Service",
        "leadRoutingEngine.js",
        "Phase 2 — Routing engine that evaluates active routing rules against inbound lead metadata and resolves the target BU and pipeline stage.",
    ),
    (
        "1.2.0",
        "30 May 2026  3:34 pm AEST",
        "101bf21",
        "Schema / Migration",
        "053_projects.sql\n054_project_task_link.sql\n055_time_logs.sql",
        "Phase 3 — Projects table (frozen baseline hours/cost, lead FK), backwards-compatible project_id column on client_work_tasks, and time-log table with cost rate and date.",
    ),
    (
        "1.2.0",
        "30 May 2026  3:34 pm AEST",
        "101bf21",
        "Backend Model",
        "Project.js\nClientWorkTask.js",
        "Phase 3 — Project CRUD with live Baseline vs Actual calculation via SQL subquery, project activity log, time log CRUD, and project-scoped task helpers.",
    ),
    (
        "1.2.0",
        "30 May 2026  3:34 pm AEST",
        "101bf21",
        "Backend Route",
        "projectRoutes.js\nplatformRouter.js",
        "Phase 3 — Project routes including the Mark-as-Won conversion bridge (SELECT FOR UPDATE transaction), time logs, activity feed, and task board endpoints.",
    ),
    (
        "1.2.0",
        "30 May 2026  3:34 pm AEST",
        "101bf21",
        "Backend Service",
        "leadConversionService.js",
        "Phase 3 — Atomic lead-to-project conversion: locks lead row, sums estimates into baseline, creates project, logs activity on both records.",
    ),
    (
        "1.2.0",
        "30 May 2026  3:39 pm AEST",
        "4148229",
        "Schema / Migration",
        "056_webhook_endpoints.sql",
        "Phase 4/5 — Webhook endpoints table with JSONB events array and dispatch log (attempt count, HTTP status, error detail).",
    ),
    (
        "1.2.0",
        "30 May 2026  3:39 pm AEST",
        "4148229",
        "Backend Model",
        "WebhookEndpoint.js",
        "Phase 4/5 — Webhook endpoint CRUD (auto-generates whsec_ signing secret), active-endpoint lookup by event name, and dispatch log writer.",
    ),
    (
        "1.2.0",
        "30 May 2026  3:39 pm AEST",
        "4148229",
        "Backend Service",
        "webhookDispatchService.js",
        "Phase 4/5 — Fire-and-forget webhook dispatcher: fetches active endpoints, HMAC-SHA256 signs payload, retries up to 3× with 2s/4s/8s backoff, 10 s AbortController timeout.",
    ),
    (
        "1.3.0",
        "30 May 2026  3:54 pm AEST",
        "6135ff6",
        "Frontend Page",
        "PlatformCRM.jsx\nPlatformLeads.jsx\nPlatformAccounts.jsx\nPlatformProjects.jsx\nPlatformProjectDetail.jsx\nPlatformCRMSettings.jsx",
        "Initial frontend build: CRM hub, Kanban pipeline board, accounts with contact management, projects list, project detail (task board + budget panel + time logs), and CRM settings (BU + webhook config).",
    ),
    (
        "1.3.0",
        "30 May 2026  3:54 pm AEST",
        "6135ff6",
        "Frontend Style",
        "crm.css",
        "CRM-specific stylesheet: pipeline board, lead cards, budget panel, status badges, hub grid, stat cards, time log list, activity feed.",
    ),
    (
        "1.3.0",
        "30 May 2026  3:54 pm AEST",
        "6135ff6",
        "Frontend Config",
        "App.jsx\nNavigation.jsx",
        "Added lazy-loaded routes for all CRM pages and a CRM nav link in the platform sidebar.",
    ),
    (
        "1.3.0",
        "30 May 2026  4:07 pm AEST",
        "61c2bd3",
        "Documentation",
        "crm-mockups.html",
        "Static HTML mockups of all five CRM screens for stakeholder review before merging.",
    ),
    (
        "1.3.0",
        "2 Jun 2026  11:53 pm AEST",
        "4a4741f",
        "Schema / Migration",
        "057_crm_tables.sql",
        "CRM rework to MVP: replaces pipeline/kanban model with user-defined schema — crm_organisations, crm_contacts, crm_notes. Business units hardcoded (Outlier Core, Outlier Skate, Rhythm Engine, Adoption Accelerator, AI-Human Workforce Design, ET Inc). Org notes and contact notes separated by CHECK constraint.",
    ),
    (
        "1.3.0",
        "2 Jun 2026  11:53 pm AEST",
        "4a4741f",
        "Backend Model",
        "CrmOrganisation.js\nCrmContact.js\nCrmNote.js",
        "CRM rework — CRUD models for organisations (with BU/lead-status filtering), contacts (nested under org), and notes (org-level and contact-level, each touching parent updated_at).",
    ),
    (
        "1.3.0",
        "2 Jun 2026  11:53 pm AEST",
        "4a4741f",
        "Backend Route",
        "crmOrgRoutes.js\nplatformRouter.js",
        "CRM rework — Full REST API at /api/platform/crm/: organisations CRUD, contact CRUD, org notes, contact notes. Mounted alongside existing routes with no impact on Rhythm Engine.",
    ),
    (
        "1.3.0",
        "2 Jun 2026  11:53 pm AEST",
        "4a4741f",
        "Frontend Page",
        "PlatformOrganisations.jsx\nPlatformOrgDetail.jsx",
        "CRM rework — Organisations list (search + BU + status filters, create modal) and org detail page (inline edit, contacts with per-contact notes, organisation notes panel).",
    ),
    (
        "1.3.0",
        "2 Jun 2026  11:53 pm AEST",
        "4a4741f",
        "Frontend Config",
        "App.jsx\nNavigation.jsx\ncrmConstants.js",
        "CRM rework — Replaced old pipeline routes with /platform/crm/organisations routes. Simplified nav to single CRM entry. Extracted BUSINESS_UNITS and LEAD_STATUSES to shared constants file.",
    ),
    (
        "1.3.0",
        "2 Jun 2026  12:03 am AEST",
        "36bfcfe",
        "Bug Fix",
        "CrmOrganisation.js\nCrmContact.js\nCrmNote.js",
        "Fixed deploy failure: models imported from non-existent db.js. Corrected to { query } from config/database.js to match the rest of the codebase.",
    ),
    (
        "1.3.0",
        "2 Jun 2026  12:19 am AEST",
        "b5902a6",
        "Bug Fix",
        "PlatformOrganisations.jsx\nPlatformOrgDetail.jsx\ncrm.css\ncrmOrgRoutes.js",
        "Three UI fixes: (1) modal missing 'card' class — added to restore opaque white background; (2) replaced platform-users-table with crm-table — centered columns, visible gridlines, distinct uppercase headers; (3) org detail now shows the real DB error instead of a blank screen.",
    ),
    (
        "1.3.0",
        "2 Jun 2026  12:27 am AEST",
        "896c732",
        "Bug Fix",
        "CrmNote.js",
        "Fixed 'column u.name does not exist' — users table uses first_name + last_name, not name. Updated both notes JOIN queries.",
    ),
    (
        "1.3.0",
        "2 Jun 2026  12:40 am AEST",
        "b71237a",
        "Frontend Page",
        "PlatformOrgDetail.jsx",
        "Notes now display full datetime (date + time, e.g. '2 Jun 2026, 10:45 am') instead of date only.",
    ),
]

# ── Column definitions ─────────────────────────────────────────────────────
columns = [
    ("Version",           14),
    ("Timestamp (AEST)",  26),
    ("Commit",            12),
    ("Category",          20),
    ("Files Changed",     46),
    ("Objective / Description", 70),
]

# ── Header row ─────────────────────────────────────────────────────────────
ws.row_dimensions[1].height = 36
for col_idx, (title, width) in enumerate(columns, start=1):
    cell = ws.cell(row=1, column=col_idx, value=title)
    cell.font      = Font(bold=True, size=11, color=WHITE)
    cell.fill      = fill(BLACK)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = border()
    ws.column_dimensions[get_column_letter(col_idx)].width = width

# ── Data rows ──────────────────────────────────────────────────────────────
for row_idx, (version, ts, commit, category, files, objective) in enumerate(rows, start=2):
    ws.row_dimensions[row_idx].height = max(15 * files.count("\n") + 22, 42)

    values = [version, ts, commit, category, files, objective]
    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.border    = border()
        cell.alignment = Alignment(vertical="top", wrap_text=True,
                                   horizontal="center" if col_idx <= 4 else "left")
        cell.font      = Font(size=10, color=BLACK)

        # Zebra stripe
        if row_idx % 2 == 0:
            cell.fill = fill(LIGHT_GREY)

        # Version column — coloured pill
        if col_idx == 1 and version in VERSION_COLOURS:
            cell.fill = fill(VERSION_COLOURS[version])
            cell.font = Font(bold=True, size=10, color=BLACK)

        # Category column — badge colour
        if col_idx == 4 and category in CATEGORY_STYLES:
            style = CATEGORY_STYLES[category]
            cell.fill = style["fill"]
            cell.font = style["font"]
            cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)

# ── Freeze header & auto-filter ────────────────────────────────────────────
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"

# ── Title block above table ────────────────────────────────────────────────
ws.insert_rows(1, 3)
ws.merge_cells("A1:F1")
title_cell = ws["A1"]
title_cell.value   = "Community Pulse — CRM Build Changelog"
title_cell.font    = Font(bold=True, size=16, color=BLACK)
title_cell.fill    = fill(ACCENT)
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 42

ws.merge_cells("A2:F2")
sub_cell = ws["A2"]
sub_cell.value   = "Branch: claude/great-shannon-imXPP  |  Repo: TheOutlierGroup/Community-Pulse  |  Generated: 9 Jun 2026"
sub_cell.font    = Font(italic=True, size=10, color=TEXT_MUTED)
sub_cell.fill    = fill(LIGHT_GREY)
sub_cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 22

ws.row_dimensions[3].height = 6  # spacer

wb.save("/home/user/Community-Pulse/CRM_Changelog.xlsx")
print("Done")
